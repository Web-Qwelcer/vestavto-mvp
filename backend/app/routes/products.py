"""
VestAvto MVP - Products Routes
"""
import io
import json
import logging
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, or_

logger = logging.getLogger(__name__)

from app.database import get_session
from app.models import Product, Category, CarModel, OrderItem, Order, OrderStatus
from app.schemas import ProductCreate, ProductUpdate, ProductResponse, UserInfo
from app.auth import get_current_user, get_current_manager, decode_token
from app.models import UserRole
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials

_security = HTTPBearer(auto_error=False)
router = APIRouter(prefix="/products", tags=["Products"])


@router.get("", response_model=List[ProductResponse])
async def get_products(
    category: Optional[Category] = None,
    car_model: Optional[CarModel] = None,
    available_only: bool = True,
    skip: int = 0,
    limit: int = 50,
    session: AsyncSession = Depends(get_session)
):
    """Отримати список товарів (публічний)"""
    query = select(Product)
    
    conditions = []
    if category:
        conditions.append(Product.category == category)
    if car_model:
        conditions.append(Product.car_model == car_model)
    if available_only:
        conditions.append(Product.is_available == True)
        conditions.append(or_(Product.is_reserved == False, Product.is_reserved.is_(None)))
    
    if conditions:
        query = query.where(and_(*conditions))
    
    query = query.order_by(Product.created_at.desc()).offset(skip).limit(limit)
    
    result = await session.execute(query)
    products = result.scalars().all()
    
    return [ProductResponse.model_validate(p) for p in products]


@router.get("/export")
async def export_products(
    token: Optional[str] = Query(None, description="JWT token (для iOS direct download)"),
    session: AsyncSession = Depends(get_session),
    credentials: HTTPAuthorizationCredentials = Depends(_security),
):
    """Вивантажити всі товари у Excel-файл (менеджер).
    Auth: Bearer header АБО ?token= query param (для iOS прямого URL)."""
    raw_token = token or (credentials.credentials if credentials else None)
    if not raw_token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    payload = decode_token(raw_token)
    if UserRole(payload.get("role", "client")) == UserRole.CLIENT:
        raise HTTPException(status_code=403, detail="Manager access required")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    result = await session.execute(
        select(Product).order_by(Product.id)
    )
    products = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Товари"

    # Header row
    headers = ["id", "name", "description", "price", "deposit",
               "category", "car_model", "is_available", "is_negotiable", "is_reserved"]
    ws.append(headers)

    # Style header
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    for col, cell in enumerate(ws[1], start=1):
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = [
            6, 35, 50, 10, 10, 16, 16, 12, 14, 12
        ][col - 1]

    # Data rows
    for p in products:
        ws.append([
            p.id,
            p.name,
            p.description or "",
            p.price,
            p.deposit,
            p.category.value if p.category else "",
            p.car_model.value if p.car_model else "",
            "true" if p.is_available else "false",
            "true" if p.is_negotiable else "false",
            "true" if p.is_reserved else "false",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=products.xlsx"}
    )


@router.post("/import")
async def import_products(
    file: UploadFile = File(...),
    session: AsyncSession = Depends(get_session),
    manager: UserInfo = Depends(get_current_manager)
):
    """Завантажити товари з Excel-файлу (менеджер). id пустий = новий, id є = оновлення."""
    from openpyxl import load_workbook

    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Файл має бути у форматі .xlsx")

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content))
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Не вдалось відкрити файл: {exc}")

    ws = wb.active
    created = 0
    updated = 0
    skipped = 0
    errors: list = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Skip completely empty rows
        if not any(cell is not None and str(cell).strip() != "" for cell in row):
            continue
        try:
            raw_id         = row[0] if len(row) > 0 else None
            name           = str(row[1] or "").strip() if len(row) > 1 else ""
            description    = str(row[2] or "").strip() if len(row) > 2 else ""
            price          = float(row[3] or 0)                         if len(row) > 3 else 0.0
            deposit        = float(row[4] or 0)                         if len(row) > 4 else 0.0
            category_val   = str(row[5] or "other").strip().lower()     if len(row) > 5 else "other"
            car_val        = str(row[6] or "other").strip().lower()     if len(row) > 6 else "other"
            available_val  = str(row[7] or "true").strip().lower()      if len(row) > 7 else "true"
            negotiable_val = str(row[8] or "false").strip().lower()     if len(row) > 8 else "false"
            reserved_val   = str(row[9] or "false").strip().lower()     if len(row) > 9 else "false"

            if not name:
                errors.append({"row": row_idx, "error": "Назва товару обов'язкова"})
                continue

            is_negotiable = negotiable_val in ("true", "1", "yes", "так")
            is_reserved   = reserved_val   in ("true", "1", "yes", "так")

            if price <= 0 and not is_negotiable:
                errors.append({"row": row_idx, "error": "Ціна має бути більше 0 (або встановіть is_negotiable=true)"})
                continue

            try:
                category = Category(category_val)
            except ValueError:
                category = Category.OTHER

            try:
                car_model = CarModel(car_val)
            except ValueError:
                car_model = CarModel.OTHER

            is_available = available_val in ("true", "1", "yes", "так")

            # Resolve product: by id → by name → create
            product = None
            if raw_id:
                res = await session.execute(
                    select(Product).where(Product.id == int(raw_id))
                )
                product = res.scalar_one_or_none()
                if not product:
                    errors.append({"row": row_idx, "error": f"Товар з id={raw_id} не знайдено"})
                    continue
            else:
                # Try to find by exact name
                res = await session.execute(
                    select(Product).where(Product.name == name).limit(1)
                )
                product = res.scalar_one_or_none()

            if product:
                changed = (
                    product.name           != name
                    or (product.description or "") != (description or "")
                    or product.price          != price
                    or product.deposit        != deposit
                    or product.category       != category
                    or product.car_model      != car_model
                    or product.is_available   != is_available
                    or product.is_negotiable  != is_negotiable
                    or product.is_reserved    != is_reserved
                )

                if not changed:
                    skipped += 1
                    continue

                product.name          = name
                product.description   = description or None
                product.price         = price
                product.deposit       = deposit
                product.category      = category
                product.car_model     = car_model
                product.is_available  = is_available
                product.is_negotiable = is_negotiable
                product.is_reserved   = is_reserved
                updated += 1
            else:
                product = Product(
                    name=name,
                    description=description or None,
                    price=price,
                    deposit=deposit,
                    category=category,
                    car_model=car_model,
                    is_available=is_available,
                    is_negotiable=is_negotiable,
                    is_reserved=is_reserved,
                )
                session.add(product)
                created += 1

        except Exception as exc:
            logger.exception(f"[Import] Row {row_idx} error: {exc}")
            errors.append({"row": row_idx, "error": str(exc)})

    await session.commit()
    return {"created": created, "updated": updated, "skipped": skipped, "errors": errors}


@router.get("/{product_id}", response_model=ProductResponse)
async def get_product(
    product_id: int,
    session: AsyncSession = Depends(get_session)
):
    """Отримати товар за ID"""
    result = await session.execute(
        select(Product).where(Product.id == product_id)
    )
    product = result.scalar_one_or_none()
    
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    return ProductResponse.model_validate(product)


@router.post("", response_model=ProductResponse)
async def create_product(
    data: ProductCreate,
    session: AsyncSession = Depends(get_session),
    manager: UserInfo = Depends(get_current_manager)
):
    """Створити товар (тільки менеджер)"""
    product = Product(
        name=data.name,
        description=data.description,
        price=data.price,
        deposit=data.deposit,
        category=data.category,
        car_model=data.car_model,
        photos=json.dumps(data.photos) if data.photos else None,
        is_available=data.is_available,
        is_negotiable=data.is_negotiable,
    )
    session.add(product)
    await session.flush()
    
    response = ProductResponse.model_validate(product)
    response.photos = data.photos or []
    return response


@router.put("/{product_id}", response_model=ProductResponse)
async def update_product(
    product_id: int,
    data: ProductUpdate,
    session: AsyncSession = Depends(get_session),
    manager: UserInfo = Depends(get_current_manager)
):
    """Оновити товар (тільки менеджер)"""
    result = await session.execute(
        select(Product).where(Product.id == product_id)
    )
    product = result.scalar_one_or_none()
    
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    update_data = data.model_dump(exclude_unset=True)
    
    if "photos" in update_data:
        update_data["photos"] = json.dumps(update_data["photos"]) if update_data["photos"] else None
    
    for field, value in update_data.items():
        setattr(product, field, value)
    
    await session.flush()
    
    return ProductResponse.model_validate(product)


@router.delete("/{product_id}")
async def delete_product(
    product_id: int,
    session: AsyncSession = Depends(get_session),
    manager: UserInfo = Depends(get_current_manager)
):
    """Видалити товар (тільки менеджер)"""
    result = await session.execute(
        select(Product).where(Product.id == product_id)
    )
    product = result.scalar_one_or_none()
    
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    # Блокуємо видалення тільки якщо товар є в АКТИВНИХ замовленнях.
    # Скасовані (cancelled) та доставлені (delivered) — не перешкода.
    active_statuses = [
        s for s in OrderStatus
        if s not in (OrderStatus.CANCELLED, OrderStatus.DELIVERED)
    ]
    linked = await session.execute(
        select(OrderItem)
        .join(Order, Order.id == OrderItem.order_id)
        .where(
            OrderItem.product_id == product_id,
            Order.status.in_(active_statuses),
        )
        .limit(1)
    )
    if linked.scalar_one_or_none():
        raise HTTPException(
            status_code=409,
            detail="Неможливо видалити товар — він є в активних замовленнях"
        )

    await session.delete(product)
    return {"ok": True}


@router.post("/{product_id}/upload-image", response_model=ProductResponse)
async def upload_product_image(
    product_id: int,
    file: UploadFile = File(...),
    session: AsyncSession = Depends(get_session),
    manager: UserInfo = Depends(get_current_manager),
):
    """Завантажити фото товару в Cloudinary (тільки менеджер)"""
    result = await session.execute(select(Product).where(Product.id == product_id))
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="File must be an image")

    file_bytes = await file.read()
    if len(file_bytes) > 10 * 1024 * 1024:  # 10 MB
        raise HTTPException(status_code=400, detail="Image too large (max 10 MB)")

    try:
        from app.services.cloudinary_service import upload_image
        url = await upload_image(file_bytes, product_id)
    except RuntimeError as exc:
        raise HTTPException(status_code=500, detail=f"Upload failed: {exc}")

    # Додаємо URL в кінець списку photos (зберігаємо порядок вибору)
    existing: list = []
    if product.photos:
        try:
            existing = json.loads(product.photos)
        except Exception:
            existing = []

    existing.append(url)
    product.photos = json.dumps(existing)
    await session.commit()

    return ProductResponse.model_validate(product)


@router.get("/categories/list")
async def get_categories():
    """Отримати список категорій"""
    return [
        {"value": c.value, "label": get_category_label(c)} 
        for c in Category
    ]


@router.get("/cars/list")
async def get_car_models():
    """Отримати список авто"""
    return [
        {"value": c.value, "label": get_car_label(c)} 
        for c in CarModel
    ]


def get_category_label(cat: Category) -> str:
    labels = {
        Category.ENGINE: "Двигун і навісне",
        Category.TRANSMISSION: "Трансмісія",
        Category.SUSPENSION: "Ходова частина",
        Category.BODY: "Кузов і оптика",
        Category.INTERIOR: "Салон",
        Category.ELECTRICAL: "Електрика",
        Category.OTHER: "Інше"
    }
    return labels.get(cat, cat.value)


def get_car_label(car: CarModel) -> str:
    labels = {
        CarModel.SUPERB_2_PRE: "Skoda Superb 2 (дорест)",
        CarModel.SUPERB_2_REST: "Skoda Superb 2 (рест)",
        CarModel.PASSAT_B7: "VW Passat B7",
        CarModel.CC: "VW CC",
        CarModel.TOUAREG: "VW Touareg",
        CarModel.TIGUAN: "VW Tiguan",
        CarModel.OTHER: "Інше"
    }
    return labels.get(car, car.value)
